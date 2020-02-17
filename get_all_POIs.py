# -*- coding: utf-8 -*-
"""
Created on Fri Feb  7 15:04:48 2020

@author: 毛征明
"""

import requests
import json
import openpyxl

#处理types数据
from openpyxl import load_workbook
wb = load_workbook(r'D:\C_research\服务外包\交通时空大数据分析挖掘系统-数据\amap_poicode.xlsx')
ws = wb['POI分类与编码（中英文）']
label=[]
try:
    for i in range(len(ws["A"])):
        if ws["C"][i].value!=ws["C"][i+1].value:
            label.append(str(ws["B"][i+1].value))
except:
    pass
label.pop()

parameters=dict()

base_url="https://restapi.amap.com/v3/place/polygon?key=3c60df6b7aa67ecdce97874711c2f085&offset=25"
parameters["polygon"]="122.3508,41.08806;124.0002823,42.50168"
parameters["page"]="1"
import re
def SD(url):
    base,coordinate=url.split("&polygon=")
    c1,c2=coordinate.split(";")
    c1=eval(c1)
    c2=eval(c2)
    ret=[]
    ret.append(base+"&polygon="+str(c1[0])+","+str(c1[1])+";"+str((c1[0]+c2[0])/2)+","+str((c1[1]+c2[1])/2))
    ret.append(base+"&polygon="+str((c1[0]+c2[0])/2)+","+str(c1[1])+";"+str(c2[0])+","+str((c1[1]+c2[1])/2))
    ret.append(base+"&polygon="+str(c1[0])+","+str((c1[1]+c2[1])/2)+";"+str((c1[0]+c2[0])/2)+","+str(c2[1]))
    ret.append(base+"&polygon="+str((c1[0]+c2[0])/2)+","+str((c1[1]+c2[1])/2)+";"+str(c2[0])+","+str(c2[1]))
    return ret
    
urls=[]
for l in label:
    flag=True
    url=base_url+"&types="+l+"&page"+"="+parameters["page"]+"&polygon"+"="+parameters["polygon"]
    unprocessed=[url]
    while unprocessed:
        url=unprocessed[0]
        ret=requests.get(url).content
        ret=json.loads(ret)
        if eval(ret["count"])<800:
            urls.append(url)
            unprocessed.remove(url)
            continue
        else:
            unprocessed.remove(url)
            unprocessed.extend(SD(url))
