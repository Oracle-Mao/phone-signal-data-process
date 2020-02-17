# -*- coding: utf-8 -*-
"""
Created on Mon Feb 17 15:58:35 2020

@author: 毛征明
"""
#读入数据准备
import json
import numpy as np
from math import sin, asin, cos, radians, fabs, sqrt
from gensim.models import LdaModel
from gensim.corpora import Dictionary
import pandas
#读入基站点序列
f=open(".\data\clean_data.csv")
clean_data=pandas.read_csv(f)
lac={}
for l in set(clean_data["lac_id"]):
    lac[l]=[]
for i in range(len(clean_data["cell_id"])):
   lac[clean_data["lac_id"][i]].append((clean_data["longitude"][i],clean_data["latitude"][i]))
for k in lac.keys():
    lon=0
    lat=0
    for l in lac[k]:
        lon+=l[0]
        lat+=l[1]
    lon/=len(lac[k])
    lat/=len(lac[k])
    lac[k]=(lon,lat)
#读POI数据
P=list(lac.values())
f=open(".\data\POIs.json")
data_all=json.load(f)

from openpyxl import load_workbook
wb = load_workbook(r'.\data\amap_poicode.xlsx')
ws = wb['POI分类与编码（中英文）']
label=[]
for i in range(len(ws["A"])-1):
    if ws["C"][i].value!=ws["C"][i+1].value:
        label.append(ws["B"][i+1].value)
label.pop()#最后一个是None
for i in range(len(label)):
    label[i]=str(label[i])
#建立对应关系，之后要做一些优化
match={}
for i in range(len(ws["A"])-1):
    match[str(ws["B"][i].value)]=ws["C"][i].value


def caculate_distance(lat0,lat1,lon0,lon1):
    lat0,lat1,lon0,lon1=map(radians,[lat0,lat1,lon0,lon1])
    dlon = fabs(lon0 - lon1)
    dlat = fabs(lat0 - lat1)
    h=sin(dlat/2)**2+cos(lat0)*cos(lat1)*sin(dlon/2)**2
    dis=2 *6371*asin(sqrt(h))#6371 is length of raidus
    return dis

def NN(coordinate,BL):
    ret=0
    nearest=6371*6.3
    for i in range(len(BL)):
        dis=caculate_distance(coordinate[1],BL[i][1],coordinate[0],BL[i][0])
        if dis<nearest:
            nearest=dis
            ret=i
    return ret

docs=[[] for _ in data_all]
for d in data_all:
    try:
        docs[NN(eval(d["coordinate"]),P)].append(match[d["type"]])
    except:
        pass

#接下来进行LDA的聚类
dictionary=Dictionary([list(match.values())])
documents=[dictionary.doc2bow(doc) for doc in docs]
ldamodel=LdaModel(documents,num_topics=4,id2word=dictionary,passes=20)
ldamodel.print_topics()
#看完结果后保存
store=[]
for i in range(len(P)):
    tem={}
    tem["coordinate"]=P[i]
    vector=ldamodel.get_document_topics(documents[i])
    tem["vector"]=vector
    for v in vector:
        if v[1]>0.4:
            tem["class"]=v[0]
    store.append(tem)

class NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        else:
            return super(NpEncoder, self).default(obj)
        

json_str=json.dumps(store, cls=NpEncoder)
with open('.\data\city_map_new.json', 'w') as json_file:
    json_file.write(json_str)
ldamodel.save(".\data\lda_model.train")
